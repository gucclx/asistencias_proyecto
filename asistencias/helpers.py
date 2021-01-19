from flask import session, redirect, url_for
from functools import wraps
from openpyxl import Workbook
import sqlite3

DB_PATH = "./asistencias/asistencias.db"

def login_requerido(f):

	"""Decora la ruta para requerir login"""

	@wraps(f)
	def wrapper():

		if session.get("user_id") is None:
			return redirect(url_for("login"))
		return f()

	return wrapper

def admin_requerido(f):

	"""Decora la ruta para requerir admin"""

	@wraps(f)
	def wrapper():

		filas = db_ejecutar("""
				SELECT admin 
				FROM Profesores 
				WHERE id = (?)""", session["user_id"])

		if not filas:
			return "Usuario inexistente.", 404

		if filas[0]["admin"] == 0:
			return "admin requerido", 403
		return f()

	return wrapper

def es_admin(prof_id):

	"""Retorna 1 o 0 dependiendo del estado de admin
		del profesor especificado.

		args:

		prof_id -- Entero. Id del profesor
	"""

	prof = db_ejecutar("""
			SELECT admin FROM Profesores 
			WHERE id = (?)""", prof_id)
	
	if not prof:
		return 0

	return prof[0]["admin"]

def db_ejecutar(query, *args, many=False):

	"""Ejecuta queries en la base de datos.

		args:

		query -- La string query a ejecutar.

		*args -- Cualquier cantidad de argumentos. Util para queries parametrizados

		kwargs:

		many (por defecto falso) -- Determina si se ejecutara el query multiples veces.
																util para realizar varios inserts a la vez.
																En caso de usar variables, *args debe ser una lista de tuplas/listas 
																con las variables requeridas para cada query
	"""

	conn = None
	filas = []
	try:
		conn = sqlite3.connect(DB_PATH)

		# configurar sqlite3 para que devuelva diccionarios en cada query
		conn.row_factory = sqlite3.Row
		cur = conn.cursor()
		cur.execute("PRAGMA foreign_keys=ON")
		
		if many:
			cur.executemany(query, args[0])
		else:
			cur.execute(query, args)

		# convertir de row object a dict
		filas = [dict(fila) for fila in cur.fetchall()]

		cur.close()

	except sqlite3.Error as e:
		print(e)
		raise Exception(e)

	finally:
		if conn:
			conn.commit()
			conn.close()

		return filas

def validar_clases(clases):

	"""Valida una lista de clases.

		args: 

		clases -- Lista de diccionarios (clases) con keys 'nombre', 'id'.

		Validaciones:

		1. El nombre debe ser string.
		2. El id de la clase debe ser int y mayor que 0.
		3. El nombre e id de la clase debe existir en la base de datos.
		4. La clase le debe pertenecer al profesor, al menos que este sea admin.

		Si alguna clase es invalida, se retorna una lista vacia.
		Si no, se retorna la misma lista
	"""
	
	clases_id = []

	try:
		for clase in clases:

			if not isinstance(clase["nombre"], str):
				raise TypeError

			clase_id = int(clase["id"])

			if clase_id <= 0:
				raise ValueError

			clases_id.append(clase["id"])

	except (TypeError, ValueError) as e:
		return []

	phs = ", ".join(["?"] * len(clases_id))

	prof_id = session["user_id"]
	admin = es_admin(prof_id)
	clases_registradas = []

	if not admin:
		clases_registradas = db_ejecutar(f"""
			SELECT Clases.nombre 
			FROM Clases
			JOIN Clase_Profesor
			ON Clases.id = Clase_profesor.clase_id
			WHERE Clases.id IN ({phs})
			AND prof_id = (?)""", *clases_id, prof_id)
	else:
		clases_registradas = db_ejecutar(f"""
			SELECT nombre 
			FROM Clases
			WHERE id IN ({phs})""", *clases_id)

	for clase in clases:
		valida = False

		for registrada in clases_registradas:

			if clase["nombre"] == registrada["nombre"]:
				valida = True
				break

		if not valida:
			return []

	return clases

def generar_excel_wb(**kwargs):

	"""Genera y retorna un workbook usando openpyxl,
		a partir de una lista de diccionarios.

		keyword args:

		titulos -- Lista de strings, estas seran los 'titulos' o 
								primeras columnas en la primera fila. No requerido.

		keys -- Lista de keys que se desean acceder en cada diccionario.

		elementos -- Lista de diccionarios.

		Cada fila en la hoja representa cada diccionario.
		Cada columna en la fila representa cada key, en el orden que aparecen.
	"""

	titulos = kwargs.get("titulos")
	keys = kwargs.get("keys")
	elementos = kwargs.get("elementos")

	wb = Workbook()
	ws = wb.active

	if titulos:
		# situar titulos
		for col in range(1, len(titulos) + 1):
			cell = ws.cell(1, col)
			cell.value = titulos[col - 1]

	# situar celdas con valores correspondientes
	for fila in range(2, len(elementos) + 2):
		elemento = elementos[fila - 2]
		for columna in range(1, len(titulos) + 1):

			celda = ws.cell(fila, columna)
			key  = keys[columna - 1]
			celda.value = elemento[key]
			
	return wb